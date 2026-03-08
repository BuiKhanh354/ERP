"""Web views (HTML pages) for the Accounting module."""
import json
from decimal import Decimal
from datetime import datetime

from django.contrib import messages
from django.contrib.auth.mixins import LoginRequiredMixin
from django.db.models import Sum, Q
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect
from django.urls import reverse_lazy
from django.utils import timezone
from django.views.generic import (
    TemplateView, ListView, DetailView, CreateView, UpdateView, DeleteView, View,
)

from core.rbac import PermissionRequiredMixin
from .models import Invoice, InvoiceItem, Payment
from .forms import InvoiceForm, InvoiceItemFormSet, PaymentForm, ExpenseFilterForm
from .services import AccountingDashboardService, ReportService
from budgeting.models import Budget, Expense, BudgetCategory
from projects.models import Project
from clients.models import Client


# ===================================================================
# Dashboard
# ===================================================================

class AccountingDashboardView(PermissionRequiredMixin, TemplateView):
    """Financial dashboard with widgets, charts, and tables."""
    template_name = 'accounting/dashboard.html'
    permission_required = 'VIEW_INVOICE'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context.update(AccountingDashboardService.get_dashboard_data())
        return context


# ===================================================================
# Invoice CRUD
# ===================================================================

class InvoiceListView(PermissionRequiredMixin, ListView):
    """List invoices with filters."""
    model = Invoice
    template_name = 'accounting/invoice_list.html'
    context_object_name = 'invoices'
    paginate_by = 20
    permission_required = 'VIEW_INVOICE'

    def get_queryset(self):
        qs = Invoice.objects.select_related('project', 'client').order_by('-issue_date')
        status = self.request.GET.get('status')
        project = self.request.GET.get('project')
        search = self.request.GET.get('search')
        if status:
            qs = qs.filter(status=status)
        if project:
            qs = qs.filter(project_id=project)
        if search:
            qs = qs.filter(
                Q(invoice_number__icontains=search) |
                Q(client__name__icontains=search)
            )
        return qs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['projects'] = Project.objects.all().order_by('name')
        context['status_choices'] = Invoice.STATUS_CHOICES
        context['current_status'] = self.request.GET.get('status', '')
        context['current_project'] = self.request.GET.get('project', '')
        context['current_search'] = self.request.GET.get('search', '')
        # Summary stats
        context['total_count'] = Invoice.objects.count()
        context['paid_count'] = Invoice.objects.filter(status='paid').count()
        context['pending_count'] = Invoice.objects.filter(status__in=['draft', 'sent']).count()
        context['overdue_count'] = Invoice.objects.filter(
            status__in=['draft', 'sent'], due_date__lt=timezone.now().date()
        ).count()
        return context


class InvoiceDetailView(PermissionRequiredMixin, DetailView):
    """View invoice details with items and payments."""
    model = Invoice
    template_name = 'accounting/invoice_detail.html'
    context_object_name = 'invoice'
    permission_required = 'VIEW_INVOICE'

    def get_queryset(self):
        return Invoice.objects.select_related('project', 'client').prefetch_related('items', 'payments')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['items'] = self.object.items.all()
        context['payments'] = self.object.payments.all().order_by('-payment_date')
        return context


class InvoiceCreateView(PermissionRequiredMixin, CreateView):
    """Create a new invoice with inline items."""
    model = Invoice
    form_class = InvoiceForm
    template_name = 'accounting/invoice_form.html'
    permission_required = 'CREATE_INVOICE'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        if self.request.POST:
            context['formset'] = InvoiceItemFormSet(self.request.POST)
        else:
            context['formset'] = InvoiceItemFormSet()
        context['title'] = 'Tạo hoá đơn mới'
        context['clients'] = Client.objects.all().order_by('name')
        return context

    def form_valid(self, form):
        context = self.get_context_data()
        formset = context['formset']
        if formset.is_valid():
            invoice = form.save(commit=False)
            invoice.created_by = self.request.user
            invoice.save()
            formset.instance = invoice
            formset.save()
            # Recalculate total
            invoice.recalculate_total()
            messages.success(self.request, f'Đã tạo hoá đơn {invoice.invoice_number} thành công.')
            return redirect('accounting:invoice_detail', pk=invoice.pk)
        else:
            return self.render_to_response(self.get_context_data(form=form))

    def get_success_url(self):
        return reverse_lazy('accounting:invoice_list')


class InvoiceUpdateView(PermissionRequiredMixin, UpdateView):
    """Edit an existing invoice."""
    model = Invoice
    form_class = InvoiceForm
    template_name = 'accounting/invoice_form.html'
    permission_required = 'EDIT_INVOICE'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        if self.request.POST:
            context['formset'] = InvoiceItemFormSet(self.request.POST, instance=self.object)
        else:
            context['formset'] = InvoiceItemFormSet(instance=self.object)
        context['title'] = f'Sửa hoá đơn {self.object.invoice_number}'
        context['clients'] = Client.objects.all().order_by('name')
        return context

    def form_valid(self, form):
        context = self.get_context_data()
        formset = context['formset']
        if formset.is_valid():
            invoice = form.save(commit=False)
            invoice.updated_by = self.request.user
            invoice.save()
            formset.save()
            invoice.recalculate_total()
            messages.success(self.request, f'Đã cập nhật hoá đơn {invoice.invoice_number}.')
            return redirect('accounting:invoice_detail', pk=invoice.pk)
        else:
            return self.render_to_response(self.get_context_data(form=form))


class InvoiceDeleteView(PermissionRequiredMixin, DeleteView):
    """Delete an invoice."""
    model = Invoice
    template_name = 'accounting/invoice_confirm_delete.html'
    success_url = reverse_lazy('accounting:invoice_list')
    permission_required = 'DELETE_INVOICE'

    def delete(self, request, *args, **kwargs):
        invoice = self.get_object()
        messages.success(request, f'Đã xoá hoá đơn {invoice.invoice_number}.')
        return super().delete(request, *args, **kwargs)


class InvoiceMarkPaidView(PermissionRequiredMixin, View):
    """Mark an invoice as paid."""
    permission_required = 'EDIT_INVOICE'

    def post(self, request, pk):
        invoice = get_object_or_404(Invoice, pk=pk)
        invoice.status = 'paid'
        invoice.updated_by = request.user
        invoice.save(update_fields=['status', 'updated_by', 'updated_at'])
        messages.success(request, f'Hoá đơn {invoice.invoice_number} đã được đánh dấu thanh toán.')
        return redirect('accounting:invoice_detail', pk=pk)


# ===================================================================
# Expense CRUD (reuses budgeting.Expense model)
# ===================================================================

class ExpenseListView(PermissionRequiredMixin, ListView):
    """List expenses with filters."""
    model = Expense
    template_name = 'accounting/expense_list.html'
    context_object_name = 'expenses'
    paginate_by = 20
    permission_required = 'VIEW_EXPENSE'

    def get_queryset(self):
        qs = Expense.objects.select_related('project', 'category').order_by('-expense_date')
        project = self.request.GET.get('project')
        category = self.request.GET.get('category')
        date_from = self.request.GET.get('date_from')
        date_to = self.request.GET.get('date_to')
        if project:
            qs = qs.filter(project_id=project)
        if category:
            qs = qs.filter(category_id=category)
        if date_from:
            qs = qs.filter(expense_date__gte=date_from)
        if date_to:
            qs = qs.filter(expense_date__lte=date_to)
        return qs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['projects'] = Project.objects.all().order_by('name')
        context['categories'] = BudgetCategory.objects.all().order_by('name')
        context['current_project'] = self.request.GET.get('project', '')
        context['current_category'] = self.request.GET.get('category', '')
        context['current_date_from'] = self.request.GET.get('date_from', '')
        context['current_date_to'] = self.request.GET.get('date_to', '')
        context['total_expenses'] = Expense.objects.aggregate(total=Sum('amount'))['total'] or 0
        return context


class ExpenseCreateView(PermissionRequiredMixin, CreateView):
    """Create a new expense."""
    model = Expense
    template_name = 'accounting/expense_form.html'
    fields = ['project', 'category', 'expense_type', 'amount', 'description', 'expense_date', 'vendor', 'invoice_number']
    success_url = reverse_lazy('accounting:expense_list')
    permission_required = 'CREATE_EXPENSE'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Tạo chi phí mới'
        context['projects'] = Project.objects.all().order_by('name')
        context['categories'] = BudgetCategory.objects.all().order_by('name')
        return context

    def form_valid(self, form):
        form.instance.created_by = self.request.user
        messages.success(self.request, 'Đã tạo chi phí thành công.')
        return super().form_valid(form)


class ExpenseUpdateView(PermissionRequiredMixin, UpdateView):
    """Edit an expense."""
    model = Expense
    template_name = 'accounting/expense_form.html'
    fields = ['project', 'category', 'expense_type', 'amount', 'description', 'expense_date', 'vendor', 'invoice_number']
    success_url = reverse_lazy('accounting:expense_list')
    permission_required = 'EDIT_EXPENSE'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = f'Sửa chi phí #{self.object.pk}'
        context['projects'] = Project.objects.all().order_by('name')
        context['categories'] = BudgetCategory.objects.all().order_by('name')
        return context

    def form_valid(self, form):
        form.instance.updated_by = self.request.user
        messages.success(self.request, 'Đã cập nhật chi phí.')
        return super().form_valid(form)


class ExpenseDeleteView(PermissionRequiredMixin, DeleteView):
    """Delete an expense."""
    model = Expense
    template_name = 'accounting/expense_confirm_delete.html'
    success_url = reverse_lazy('accounting:expense_list')
    permission_required = 'DELETE_EXPENSE'

    def delete(self, request, *args, **kwargs):
        messages.success(request, 'Đã xoá chi phí.')
        return super().delete(request, *args, **kwargs)


# ===================================================================
# Payment CRUD
# ===================================================================

class PaymentListView(PermissionRequiredMixin, ListView):
    """List payments."""
    model = Payment
    template_name = 'accounting/payment_list.html'
    context_object_name = 'payments'
    paginate_by = 20
    permission_required = 'VIEW_PAYMENT'

    def get_queryset(self):
        qs = Payment.objects.select_related(
            'invoice', 'invoice__client', 'invoice__project'
        ).order_by('-payment_date')
        invoice = self.request.GET.get('invoice')
        method = self.request.GET.get('method')
        if invoice:
            qs = qs.filter(invoice_id=invoice)
        if method:
            qs = qs.filter(payment_method=method)
        return qs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['invoices'] = Invoice.objects.all().order_by('-issue_date')
        context['method_choices'] = Payment.PAYMENT_METHOD_CHOICES
        context['current_invoice'] = self.request.GET.get('invoice', '')
        context['current_method'] = self.request.GET.get('method', '')
        context['total_payments'] = Payment.objects.aggregate(total=Sum('amount'))['total'] or 0
        return context


class PaymentCreateView(PermissionRequiredMixin, CreateView):
    """Record a new payment."""
    model = Payment
    form_class = PaymentForm
    template_name = 'accounting/payment_form.html'
    success_url = reverse_lazy('accounting:payment_list')
    permission_required = 'CREATE_PAYMENT'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Ghi nhận thanh toán'
        # Pre-select invoice if provided in URL
        invoice_id = self.request.GET.get('invoice')
        if invoice_id:
            context['preselected_invoice'] = invoice_id
        return context

    def form_valid(self, form):
        form.instance.created_by = self.request.user
        payment = form.save()
        # Check if invoice is fully paid
        invoice = payment.invoice
        if invoice.amount_due <= 0:
            invoice.status = 'paid'
            invoice.save(update_fields=['status', 'updated_at'])
        messages.success(self.request, f'Đã ghi nhận thanh toán {payment.amount:,.0f}đ.')
        return redirect(self.success_url)


class PaymentUpdateView(PermissionRequiredMixin, UpdateView):
    """Edit a payment."""
    model = Payment
    form_class = PaymentForm
    template_name = 'accounting/payment_form.html'
    success_url = reverse_lazy('accounting:payment_list')
    permission_required = 'EDIT_PAYMENT'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = f'Sửa thanh toán #{self.object.pk}'
        return context

    def form_valid(self, form):
        form.instance.updated_by = self.request.user
        messages.success(self.request, 'Đã cập nhật thanh toán.')
        return super().form_valid(form)


# ===================================================================
# Budget Summary
# ===================================================================

class BudgetSummaryView(PermissionRequiredMixin, TemplateView):
    """Budget summary with Budget vs Actual chart."""
    template_name = 'accounting/budget_summary.html'
    permission_required = 'VIEW_BUDGET'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        budgets = Budget.objects.select_related('project', 'category').order_by('project__name', 'category__name')
        context['budgets'] = budgets

        # Chart data: Budget vs Actual per project
        budget_by_project = Budget.objects.values('project__name').annotate(
            total_budget=Sum('allocated_amount'),
            total_spent=Sum('spent_amount'),
        ).order_by('project__name')

        context['budget_chart_data'] = json.dumps({
            'labels': [b['project__name'] for b in budget_by_project],
            'budget': [float(b['total_budget']) for b in budget_by_project],
            'actual': [float(b['total_spent']) for b in budget_by_project],
        })

        # Summary totals
        totals = Budget.objects.aggregate(
            total_budget=Sum('allocated_amount'),
            total_spent=Sum('spent_amount'),
        )
        context['total_budget'] = totals['total_budget'] or 0
        context['total_spent'] = totals['total_spent'] or 0
        context['total_remaining'] = (totals['total_budget'] or 0) - (totals['total_spent'] or 0)
        return context


# ===================================================================
# Reports
# ===================================================================

class ReportListView(PermissionRequiredMixin, TemplateView):
    """Report selection page."""
    template_name = 'accounting/reports.html'
    permission_required = 'VIEW_REPORT'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['projects'] = Project.objects.all().order_by('name')
        return context


class ReportExportView(PermissionRequiredMixin, View):
    """Export reports as Excel."""
    permission_required = 'VIEW_REPORT'

    def get(self, request):
        report_type = request.GET.get('type', 'profit_loss')
        date_from = request.GET.get('date_from')
        date_to = request.GET.get('date_to')
        project_id = request.GET.get('project')

        # Parse dates
        df = None
        dt = None
        if date_from:
            try:
                df = datetime.strptime(date_from, '%Y-%m-%d').date()
            except ValueError:
                pass
        if date_to:
            try:
                dt = datetime.strptime(date_to, '%Y-%m-%d').date()
            except ValueError:
                pass

        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        except ImportError:
            return HttpResponse('openpyxl chưa được cài đặt. Vui lòng chạy: pip install openpyxl', status=500)

        wb = openpyxl.Workbook()
        ws = wb.active

        header_font = Font(bold=True, size=12, color='FFFFFF')
        header_fill = PatternFill(start_color='0000CD', end_color='0000CD', fill_type='solid')
        header_alignment = Alignment(horizontal='center')
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin'),
        )

        if report_type == 'profit_loss':
            ws.title = 'Lãi Lỗ theo Dự án'
            data = ReportService.project_profit_loss(project_id, df, dt)
            headers = ['Dự án', 'Doanh thu', 'Chi phí', 'Lãi/Lỗ']
            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=h)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border
            for row_idx, item in enumerate(data, 2):
                ws.cell(row=row_idx, column=1, value=item['project']).border = thin_border
                ws.cell(row=row_idx, column=2, value=float(item['revenue'])).border = thin_border
                ws.cell(row=row_idx, column=3, value=float(item['expenses'])).border = thin_border
                ws.cell(row=row_idx, column=4, value=float(item['profit_loss'])).border = thin_border
            for col in range(1, 5):
                ws.column_dimensions[chr(64 + col)].width = 25

        elif report_type == 'monthly_revenue':
            ws.title = 'Doanh thu theo tháng'
            data = ReportService.monthly_revenue(df, dt)
            headers = ['Tháng', 'Doanh thu', 'Số hoá đơn']
            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=h)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border
            for row_idx, item in enumerate(data, 2):
                ws.cell(row=row_idx, column=1, value=item['month']).border = thin_border
                ws.cell(row=row_idx, column=2, value=float(item['total'])).border = thin_border
                ws.cell(row=row_idx, column=3, value=item['count']).border = thin_border
            for col in range(1, 4):
                ws.column_dimensions[chr(64 + col)].width = 25

        elif report_type == 'expense_category':
            ws.title = 'Chi phí theo danh mục'
            data = ReportService.expense_by_category(df, dt)
            headers = ['Danh mục', 'Tổng chi phí', 'Số mục']
            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=h)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border
            for row_idx, item in enumerate(data, 2):
                ws.cell(row=row_idx, column=1, value=item['category__name']).border = thin_border
                ws.cell(row=row_idx, column=2, value=float(item['total'])).border = thin_border
                ws.cell(row=row_idx, column=3, value=item['count']).border = thin_border
            for col in range(1, 4):
                ws.column_dimensions[chr(64 + col)].width = 25

        elif report_type == 'cash_flow':
            ws.title = 'Dòng tiền'
            data = ReportService.cash_flow(df, dt)
            headers = ['Tháng', 'Doanh thu', 'Chi phí', 'Dòng tiền ròng']
            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=h)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border
            for row_idx, item in enumerate(data, 2):
                ws.cell(row=row_idx, column=1, value=item['month']).border = thin_border
                ws.cell(row=row_idx, column=2, value=item['revenue']).border = thin_border
                ws.cell(row=row_idx, column=3, value=item['expenses']).border = thin_border
                ws.cell(row=row_idx, column=4, value=item['net_cash_flow']).border = thin_border
            for col in range(1, 5):
                ws.column_dimensions[chr(64 + col)].width = 25
        else:
            return HttpResponse('Loại báo cáo không hợp lệ.', status=400)

        # Build response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f'report_{report_type}_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response
